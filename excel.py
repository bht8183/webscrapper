import pandas as pd
from openpyxl import load_workbook
from details import extract_company_website
from openpyxl.styles import Alignment

def create_excel():
    # Define the columns as specified
    columns = [
        "ID", "Posted Date", "Company", "Website", "HTSUS Code", "Product Class",
        "Previously Granted Exclusion ID", "Total Annual Exclusion \nRequest KG",
        "3 year average (2015-2017) \nannual consumption KG", "Product Information",
        "Product Commercial Names", "Product Association Code", "Product Application"
    ]

    # Create an empty DataFrame with the specified columns
    df = pd.DataFrame(columns=columns)

    # Define the Excel file name
    excel_file = "output.xlsx"

    # Save the DataFrame to an Excel file
    df.to_excel(excel_file, index=False, engine='openpyxl')

    # Load the workbook and select the active worksheet
    wb = load_workbook(excel_file)
    ws = wb.active

    # Define column widths (example widths, you can adjust them as needed)
    column_widths = {
        'A': 10,  # ID
        'B': 15,  # Posted Date
        'C': 20,  # Company
        'D': 25,  # Website
        'E': 15,  # HTSUS Code
        'F': 15,  # Product Class
        'G': 30,  # Previously Granted Exclusion ID
        'H': 25,  # Total Annual Exclusion Request KG
        'I': 30,  # 3 year average (2015-2017) annual consumption KG
        'J': 40,  # Product Information
        'K': 25,  # Product Commercial Names
        'L': 25,  # Product Association Code
        'M': 50   # Product Application
    }

    # Set the width for each column
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Center align the text in all cells
    alignment = Alignment(horizontal='center', vertical='center')
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = alignment

    # Double the height of the first row
    ws.row_dimensions[1].height = ws.row_dimensions[1].height * 2 if ws.row_dimensions[1].height else 30

    # Save the workbook with the adjusted column widths
    wb.save(excel_file)

    print(f"Excel file '{excel_file}' created successfully with specified column widths.")

def insert_data_into_excel(data, excel_file="output.xlsx"):
    # Load the existing workbook and sheet
    wb = load_workbook(excel_file)
    ws = wb.active

    # Define the alignment for centering text
    alignment = Alignment(horizontal='center', vertical='center')

    # Get the current maximum row number
    max_row = ws.max_row
    
    # Loop through each record and insert into the worksheet
    for record in data:
        id = record[0]
        details = extract_company_website(id)

        row = [
            id,
            record[6],  # Posted Date
            record[1],  # Company
            details['company_website'],
            record[3],  # HTSUS Code
            details['metal_class'],
            details['previously_granted_er'],
            details['total_requested_annual'],
            details['avg_annual_consumption'],
            details['product_description'],
            ", ".join(details['commercial_names']) if details['commercial_names'] else "",
            "",  # Product Association Code
            details['product_applications']
        ]

        # Append the row to the worksheet
        ws.append(row)

        max_row += 1

        # Center align the newly added row
        for cell in ws[max_row]:
            cell.alignment = alignment

    # Save the workbook
    wb.save(excel_file)
    print(f"Data inserted into '{excel_file}' successfully.")